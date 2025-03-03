import datetime
import json
import re
import time

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

import proxy_gpt


def get_test_data():
    with open("summary-testing/test_data.json", "r") as f:
        data = json.loads(f.read())

    res = [item['input'] + item['output'] for item in data if item['name'] == "dialog_personal_context"]
    return res


def get_interview_from_file(path: str):
    with open(path, "r") as f:
        data = f.read()

    return data


def get_question_score(question: str):
    prompt = 'Предложение для оценки: {}\n' \
             'Оцени, является ли предложение вопросом. Где 0 - не вопрос, 1 - вопрос.\n' \
             'В ответе пришли только оценку в виде числа'
    return int(proxy_gpt.send_request([prompt.format(question)]))


def get_answers(dialog: str, questions: list[str]):
    prompt = 'Вопросы: {}\n' \
             'Ответь на присланные мной вопросы информацией из текста. Требование к ответу на вопрос:\n' \
             '1. Если ответа нет, то в качестве ответа пришли пустую строку\n' \
             '2. Если в вопросе содержится несколько подвопросов, то ответь на каждый из них и соедини ответы в единый текст' \
             'В ответе пришли json массив из строк. Формат: [\"{{ОтветНа1Вопрос}}\", \"{{ОтветНа2Вопрос}}\"...]. ' \
             'Ответ на каждый вопрос - отдельный элемент массива. ' \
             'Массив должен быть упорядочен: ответы пришли в том же порядке, как были присланы вопросы. ' \
             'Количество элементов массива должно быть = {}. Каждый элемент - строка, ограниченная двойными кавычками \"\n' \
             'Пример ответа на запрос [\"Сколько тебе лет\", \"Как тебя зовут?\"]: [\"17\", \"Лёша\"]\n'

    print(questions)

    answers = json.loads(proxy_gpt.send_request([dialog, prompt.format(questions, len(questions))]).replace('\n', ''))

    return answers


def get_questions_from_interview(interview: str):
    phrase_template = r'Speaker \d*: (.+)'
    dialog_phrases = interview.split('\n\n')

    questions = []
    for phrase in dialog_phrases:
        potential_question = re.search(phrase_template, phrase).group(1)
        while True:
            try:
                if len(potential_question) <= 10:
                    break
                time.sleep(1)
                question_score = get_question_score(potential_question)
                if question_score >= 1:
                    questions.append(potential_question)
                break
            except:
                pass
    return questions


def is_question(potential_question: str):
    while True:
        try:
            return get_question_score(potential_question) == 1
        except Exception as e:
            print(e)


def reduce_answer(question: str, answer: str):
    prompt = 'Перед тобой вопрос и ответ, который был на него дан. Сократи ответ. ' \
             'Сокращенный ответ должен быть не более 30 слов. В ответе верни сокращенный вариант ответа\n' \
             'Вопрос: {}\n' \
             'Ответ: {}'

    print("Сокращение ответа: {}".format(answer))

    return proxy_gpt.send_request([prompt.format(question, answer)])


def is_answer_correct(answers: list[str]):
    if len(answers) == 0 or len(answers[0]) == 0:
        return False

    answer = '\n'.join(answers)
    if "В тексте рассказано" in answer:
        return False
    if "В тексте нет информации" in answer:
        return False
    if "Нет информации в тексте" in answer:
        return False
    if "Пустая строка" in answer:
        return False

    return True



def get_questions_answers_pairs(interview: str, interviewer_id="0"):
    phrase_template = r'Speaker (\d*): (.+)'
    dialog_phrases = interview.split('\n\n')

    questions_answers_pairs = []
    for i, phrase in enumerate(dialog_phrases):
        speaker_id = re.search(phrase_template, phrase).group(1)
        potential_question = re.search(phrase_template, phrase).group(2)

        if speaker_id != interviewer_id:
            continue

        if len(potential_question) <= 10:
            continue

        if is_question(potential_question):
            part_of_interview = phrase
            for j in range(i + 1, min(i + 5, len(dialog_phrases))):
                phrase_additional = dialog_phrases[j]
                if len(part_of_interview) + len(phrase_additional) >= 8000:
                    break
                else:
                    part_of_interview = part_of_interview + "\n\n" + phrase_additional

            print("Текст кусочка интервью с вопросом: {}".format(part_of_interview))
            print("Длина текста: {}\n\n".format(len(part_of_interview)))

            answers = ["Ошибка при формировании ответа на вопрос"]

            for _ in range(6):
                try:
                    answers = get_answers(part_of_interview, [potential_question])

                    if is_answer_correct(answers):
                        break
                    else: continue

                except Exception as e:
                    print(e)

            #Если даже после многочисленных прогонов выдал результат нулевой длины
            if not is_answer_correct(answers):
                continue

            answer = '\n'.join(answers)
            if len(answer.split(' ')) > 40:
                answer = reduce_answer(potential_question, answer)

            questions_answers_pairs.append((potential_question, answer))

            print(questions_answers_pairs)

    return questions_answers_pairs


def create_excel_with_wrapped_text(file_name, columns):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.append(["Вопрос", "Ответ"])
    
    for row_idx, (col1_text, col2_text) in enumerate(columns, start=2):
        cell = ws.cell(row=row_idx, column=1, value=col1_text)
        cell.alignment = Alignment(wrapText=True)
        cell = ws.cell(row=row_idx, column=2, value=col2_text)
        cell.alignment = Alignment(wrapText=True)
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 80
    
    wb.save(file_name)


def main():
    # for i in range(6, 7):
    start_time = datetime.datetime.now()
    file_name = "никита_экопси.txt"
    data = get_interview_from_file("texts-for-summary/{}".format(file_name))

    questions_answers = get_questions_answers_pairs(data, '1')

    print("Время обработки интервью : {}".format(datetime.datetime.now() - start_time))
    df = pd.DataFrame({
        'Вопрос': [i[0] for i in questions_answers],
        'Ответ': [i[1] for i in questions_answers],
        'Оценка ответа': [''] * len(questions_answers)})
    df.to_excel('summary-results/Суммаризация_{}.xlsx'.format(file_name), index=False)


if __name__ == "__main__":
    main()
