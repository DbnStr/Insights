import proxy_gpt
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from templates import templates
def generate_report(template_type, text):
    base_command = 'Ответь кратко, по делу и без воды. Не выделяй ничего жирным. Ответ ищи в тексту, который я тебе ранее отправил.'
    template = templates[template_type]
    parameters = template['parameters']
    descriptions = template['parameters_description']
    result = []

    for i, desc in enumerate(descriptions):
        messages = [text, f'{desc} {base_command}']
        answer = proxy_gpt.send_request(messages)
        result.append((parameters[i], answer))
    
    return result


def create_excel_with_wrapped_text(file_name, columns):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.append(["Параметр отчета", "Значение"])
    
    for row_idx, (col1_text, col2_text) in enumerate(columns, start=2):
        ws.cell(row=row_idx, column=1, value=col1_text)
        cell = ws.cell(row=row_idx, column=2, value=col2_text)
        cell.alignment = Alignment(wrapText=True)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    
    wb.save(file_name)

# with open("example.txt", "r", encoding="utf-8") as file:
#     text = file.read()
#     generate_report('meeting', text)